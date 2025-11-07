# excel_writer.py — dynamic-column Excel writer using openpyxl only
# Unchanged logic; just extra comments for clarity.
# - Creates the workbook if it doesn't exist.
# - Creates sheets on demand.
# - Adds new headers the first time a new field appears.
# - Appends rows without enforcing a fixed schema.

from pathlib import Path
from typing import Dict, List, Iterable
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# --------- config: sheet names ----------
SHEET_COOKIE_COMPARISON = "Cookie Field Comparison"
SHEET_CLEAN_DATA = "Clean_Data"
SHEET_DIAGNOSTICS = "Diagnostics"

# --------- helpers ----------

def _open_or_create(path: Path):
    """Open workbook if present, otherwise create a fresh one."""
    path = Path(path)
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        # The default sheet will be reused/renamed on first write.
    return wb

def _ensure_sheet(wb, sheet_name: str) -> Worksheet:
    """Return an existing sheet or create/rename an empty default one."""
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    # If the default "Sheet" is empty, reuse it by renaming.
    if (
        len(wb.sheetnames) == 1 and
        wb.active.max_row == 1 and wb.active.max_column == 1 and
        wb.active["A1"].value is None
    ):
        ws = wb.active
        ws.title = sheet_name
        return ws
    return wb.create_sheet(title=sheet_name)

def _header_map(ws: Worksheet) -> Dict[str, int]:
    """
    Return a mapping of header -> 1-based column index.
    If no header row yet, returns {}.
    """
    if ws.max_row < 1:
        return {}
    headers = {}
    row = ws[1]
    for idx, cell in enumerate(row, start=1):
        val = cell.value
        if isinstance(val, str) and val.strip() != "":
            headers[val] = idx
    return headers

def _ensure_headers(ws: Worksheet, needed_headers: Iterable[str]) -> Dict[str, int]:
    """
    Make sure every header in needed_headers exists. Missing ones are appended in row 1.
    Return the up-to-date header->column map.
    """
    hdrs = _header_map(ws)
    if ws.max_row < 1 or not hdrs:
        # Initialize with the requested headers (ordered)
        for col_idx, key in enumerate(needed_headers, start=1):
            ws.cell(row=1, column=col_idx, value=key)
        return _header_map(ws)

    # Append any missing headers at the end
    next_col = ws.max_column + 1
    added = False
    for key in needed_headers:
        if key not in hdrs:
            ws.cell(row=1, column=next_col, value=key)
            hdrs[key] = next_col
            next_col += 1
            added = True
    return _header_map(ws) if added else hdrs

def _append_row(ws: Worksheet, row_dict: Dict[str, object]):
    """Append a row; add any missing headers on the fly."""
    headers_needed = list(row_dict.keys())
    hdr_map = _ensure_headers(ws, headers_needed)
    r = ws.max_row + 1 if ws.max_row >= 1 else 1
    for key, val in row_dict.items():
        c = hdr_map[key]
        ws.cell(row=r, column=c, value=val)

# --------- public API ----------

def append_cookie_comparison(out_workbook: Path, wide_row: Dict[str, object]):
    """Write a single 'wide' comparison row to 'Cookie Field Comparison'."""
    wb = _open_or_create(out_workbook)
    ws = _ensure_sheet(wb, SHEET_COOKIE_COMPARISON)
    _append_row(ws, wide_row)
    wb.save(out_workbook)

def append_clean_data_row(_master_workbook: Path, out_workbook: Path, clean_row: Dict[str, object]):
    """
    Append a row to 'Clean_Data'. We don't enforce a master schema; headers appear as needed.
    (master_workbook is accepted for backward-compat only.)
    """
    wb = _open_or_create(out_workbook)
    ws = _ensure_sheet(wb, SHEET_CLEAN_DATA)
    _append_row(ws, clean_row)
    wb.save(out_workbook)

def append_diagnostics(out_workbook: Path, rows: List[Dict[str, object]]):
    """
    Append multiple rows to 'Diagnostics'.
    Ensures the union of keys across rows exists as headers before writing.
    """
    if not rows:
        return
    wb = _open_or_create(out_workbook)
    ws = _ensure_sheet(wb, SHEET_DIAGNOSTICS)

    # Ensure union headers exist first (for stable columns)
    union_keys = []
    seen = set()
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                union_keys.append(k)
    _ensure_headers(ws, union_keys)

    for r in rows:
        _append_row(ws, r)

    wb.save(out_workbook)
